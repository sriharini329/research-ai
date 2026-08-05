import json
import os
import shutil
import zipfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# File Paths
ANDROID_REPORT_PATH = "appium/mochawesome-report/mochawesome-android.json"
WEB_REPORT_PATH = "selenium/mochawesome-report/mochawesome-web.json"
OUTPUT_DIR = "E2E-Automation-Reports"
ZIP_NAME = "E2E-Automation-Reports.zip"

def parse_mochawesome(file_path, module_prefix, type_label):
    tests = []
    total_passed = 0
    total_failed = 0
    
    if not os.path.exists(file_path):
        return total_passed, total_failed, tests

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        suites = data.get('results', [])
        idx = 1
        for suite in suites:
            for s in suite.get('suites', []):
                for test in s.get('tests', []):
                    title = test.get('title', 'Unknown Test')
                    status = "Fail"
                    if test.get('pass'):
                        status = "Pass"
                        total_passed += 1
                    elif test.get('fail'):
                        status = "Fail"
                        total_failed += 1
                    else:
                        status = "Skip"

                    # Parse TC_WEB_001 - [Feature] Description format
                    tc_id = f"TC-{module_prefix}-{idx:02d}"
                    feature = "General"
                    desc = title
                    
                    if " - " in title:
                        parts = title.split(" - ", 1)
                        if parts[0].startswith("TC_"):
                            tc_id = parts[0]
                            desc = parts[1]
                            
                    if desc.startswith("[") and "]" in desc:
                        f_end = desc.find("]")
                        feature = desc[1:f_end].strip()
                        desc = desc[f_end+1:].strip()

                    err = test.get('err', {})
                    err_details = err.get('message', '') if err else ''
                    screenshot_path = ""
                    if status == "Fail":
                        safe_name = title.replace(' ', '_').replace('/', '_')
                        screenshot_path = f"selenium/screenshots/fail_{safe_name}.png"
                    
                    tests.append({
                        'ID': tc_id,
                        'Module': type_label,
                        'Feature': feature,
                        'Description': desc,
                        'Type': type_label,
                        'Expected': 'Test should execute and complete successfully.',
                        'Actual': 'Completed successfully without any assertions failing.' if status == 'Pass' else err_details,
                        'Status': status,
                        'Duration': f"{test.get('duration', 0)} ms",
                        'Error': err_details,
                        'Screenshot': screenshot_path,
                        'Date': datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.000Z')
                    })
                    idx += 1
    except Exception as e:
        print(f"Error parsing {file_path}: {e}")

    return total_passed, total_failed, tests

def generate_excel(tests, total_tests, passed, failed, pass_rate, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "E2E Test Report"
    
    # Styles
    title_font = Font(name="Arial", size=16, color="FFFFFFFF", bold=True)
    title_fill = PatternFill(start_color="FF2E7D32", end_color="FF2E7D32", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    header_font = Font(name="Arial", size=10, color="FFFFFFFF", bold=True)
    header_fill = PatternFill(start_color="FF1B5E20", end_color="FF1B5E20", fill_type="solid")
    
    border = Border(
        left=Side(style='thin', color='FFCCCCCC'),
        right=Side(style='thin', color='FFCCCCCC'),
        top=Side(style='thin', color='FFCCCCCC'),
        bottom=Side(style='thin', color='FFCCCCCC')
    )
    
    # Header A1
    ws.merge_cells('A1:L1')
    ws['A1'] = 'End-to-End Test Automation Execution Report'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center_align
    
    # Summary Table
    ws['A3'] = 'Execution Summary'
    ws['A3'].font = Font(bold=True)
    
    ws['A4'] = 'Total Tests'
    ws['B4'] = total_tests
    ws['D4'] = 'Failed'
    ws['E4'] = failed
    
    ws['A5'] = 'Passed'
    ws['B5'] = passed
    ws['D5'] = 'Pass Rate'
    ws['E5'] = pass_rate
    
    # Column Headers
    headers = [
        'Test Case ID', 'Module', 'Feature', 'Test Description', 'Test Type', 
        'Expected Result', 'Actual Result', 'Status', 'Execution Time', 
        'Error Details', 'Screenshot Path', 'Execution Date'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        
    # Column Widths
    col_widths = {
        'A': 16.0, 'B': 14.0, 'C': 29.0, 'D': 40.0, 'E': 14.0, 'F': 40.0, 
        'G': 13.0, 'H': 12.0, 'I': 18.0, 'J': 17.0, 'K': 19.0, 'L': 28.0
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
        
    # Data Rows
    row_idx = 8
    for test in tests:
        row_data = [
            test['ID'], test['Module'], test['Feature'], test['Description'], test['Type'],
            test['Expected'], test['Actual'], test['Status'], test['Duration'],
            test['Error'], test['Screenshot'], test['Date']
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.border = border
            if col_idx == 8: # Status color
                if val == 'Pass':
                    cell.font = Font(color="FF10B981")
                elif val == 'Fail':
                    cell.font = Font(color="FFF43F5E")
        row_idx += 1
        
    wb.save(output_path)

def generate_html(tests, total_tests, passed, failed, pass_rate, output_path):
    html_template = f"""
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <title>Test Automation Dashboard</title>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
  <style>
    :root {{
      --bg-dark: #0f172a;
      --bg-card: #1e293b;
      --border: #334155;
      --text-main: #f8fafc;
      --text-muted: #94a3b8;
      --green: #10b981;
      --red: #f43f5e;
      --blue: #3b82f6;
    }}
    
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ font-family: 'Inter', sans-serif; background-color: var(--bg-dark); color: var(--text-main); padding: 2rem; min-height: 100vh; }}
    .header {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 2rem; border-bottom: 1px solid var(--border); padding-bottom: 1.5rem; }}
    .header h1 {{ font-size: 1.75rem; font-weight: 700; background: linear-gradient(to right, #10b981, #3b82f6); -webkit-background-clip: text; -webkit-text-fill-color: transparent; }}
    .meta-info {{ text-align: right; color: var(--text-muted); font-size: 0.875rem; }}
    .stats-container {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 1.5rem; margin-bottom: 2.5rem; }}
    .stat-card {{ background-color: var(--bg-card); border: 1px solid var(--border); border-radius: 12px; padding: 1.5rem; text-align: center; box-shadow: 0 4px 6px -1px rgb(0 0 0 / 0.1); }}
    .stat-label {{ font-size: 0.875rem; color: var(--text-muted); margin-bottom: 0.5rem; text-transform: uppercase; letter-spacing: 0.05em; }}
    .stat-value {{ font-size: 2.25rem; font-weight: 700; }}
    .stat-card.passed .stat-value {{ color: var(--green); }}
    .stat-card.failed .stat-value {{ color: var(--red); }}
    .stat-card.rate .stat-value {{ background: linear-gradient(to right, var(--green), #6ee7b7); -webkit-background-clip: text; -webkit-text-fill-color: transparent; }}
    .filter-bar {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 1.5rem; }}
    .filter-buttons {{ display: flex; gap: 0.75rem; }}
    .filter-btn {{ background-color: var(--bg-card); border: 1px solid var(--border); color: var(--text-main); padding: 0.5rem 1rem; border-radius: 8px; font-weight: 500; cursor: pointer; font-size: 0.875rem; }}
    .filter-btn.active {{ background-color: var(--blue); border-color: var(--blue); }}
    .search-input {{ background-color: var(--bg-card); border: 1px solid var(--border); color: var(--text-main); padding: 0.5rem 1rem; border-radius: 8px; font-size: 0.875rem; width: 280px; outline: none; }}
    .table-container {{ background-color: var(--bg-card); border: 1px solid var(--border); border-radius: 12px; overflow: hidden; box-shadow: 0 10px 15px -3px rgb(0 0 0 / 0.1); }}
    table {{ width: 100%; border-collapse: collapse; text-align: left; font-size: 0.95rem; }}
    th {{ background-color: #0f172a; color: var(--text-muted); font-weight: 600; padding: 1rem 1.5rem; border-bottom: 1px solid var(--border); text-transform: uppercase; font-size: 0.75rem; letter-spacing: 0.05em; }}
    td {{ padding: 1.25rem 1.5rem; border-bottom: 1px solid var(--border); vertical-align: middle; }}
    .test-row {{ cursor: pointer; transition: background-color 0.15s; }}
    .test-row:hover {{ background-color: #334155; }}
    .description-cell {{ display: flex; flex-direction: column; gap: 0.25rem; }}
    .type-badge {{ font-size: 0.75rem; color: var(--text-muted); background-color: #0f172a; padding: 0.1rem 0.4rem; border-radius: 4px; align-self: flex-start; }}
    .status-badge {{ display: inline-block; padding: 0.25rem 0.75rem; border-radius: 9999px; font-weight: 600; font-size: 0.75rem; text-transform: uppercase; letter-spacing: 0.05em; }}
    .status-pass {{ background-color: rgba(16, 185, 129, 0.15); color: var(--green); }}
    .status-fail {{ background-color: rgba(244, 63, 94, 0.15); color: var(--red); }}
    .time-cell {{ color: var(--text-muted); font-size: 0.875rem; }}
  </style>
</head>
<body>
  <div class="header">
    <div>
      <h1>Yoga App Test Automation Dashboard</h1>
      <p style="color: var(--text-muted); margin-top: 0.25rem;">Automated Selenium E2E Diagnostics</p>
    </div>
    <div class="meta-info">
      <div>Execution Date: <strong>{datetime.utcnow().strftime('%m/%d/%Y, %I:%M:%S %p')}</strong></div>
    </div>
  </div>

  <div class="stats-container">
    <div class="stat-card">
      <div class="stat-label">Total Tests</div>
      <div class="stat-value">{total_tests}</div>
    </div>
    <div class="stat-card passed">
      <div class="stat-label">Passed</div>
      <div class="stat-value">{passed}</div>
    </div>
    <div class="stat-card failed">
      <div class="stat-label">Failed</div>
      <div class="stat-value">{failed}</div>
    </div>
    <div class="stat-card rate">
      <div class="stat-label">Pass Rate</div>
      <div class="stat-value">{pass_rate}</div>
    </div>
  </div>

  <div class="table-container">
    <table id="test-table">
      <thead>
        <tr>
          <th style="width: 120px;">ID</th>
          <th style="width: 140px;">Module</th>
          <th style="width: 160px;">Feature</th>
          <th>Description</th>
          <th style="width: 120px;">Status</th>
          <th style="width: 120px;">Duration</th>
        </tr>
      </thead>
      <tbody>
"""
    for idx, test in enumerate(tests):
        status_class = "pass" if test['Status'] == 'Pass' else "fail"
        html_template += f"""
      <tr class="test-row {status_class}">
        <td>{test['ID']}</td>
        <td>{test['Module']}</td>
        <td>{test['Feature']}</td>
        <td class="description-cell">
          <strong>{test['Description']}</strong>
          <span class="type-badge">{test['Type']}</span>
        </td>
        <td><span class="status-badge status-{status_class}">{test['Status']}</span></td>
        <td class="time-cell">{test['Duration']}</td>
      </tr>
"""
    html_template += """
      </tbody>
    </table>
  </div>
</body>
</html>
"""
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_template)

def main():
    print("Generating Multi-Tier Enterprise E2E Report...")
    
    and_p, and_f, and_tests = parse_mochawesome(ANDROID_REPORT_PATH, "E2E", "Appium")
    web_p, web_f, web_tests = parse_mochawesome(WEB_REPORT_PATH, "WEB", "Selenium")
    
    all_tests = and_tests + web_tests
    total_tests = len(all_tests)
    passed = and_p + web_p
    failed = and_f + web_f
    pass_rate = f"{(passed / total_tests * 100):.1f}%" if total_tests > 0 else "0.0%"
    
    if os.path.exists(OUTPUT_DIR):
        shutil.rmtree(OUTPUT_DIR)
        
    os.makedirs(f"{OUTPUT_DIR}/reports", exist_ok=True)
    os.makedirs(f"{OUTPUT_DIR}/selenium/reports", exist_ok=True)
    os.makedirs(f"{OUTPUT_DIR}/selenium/screenshots", exist_ok=True)
    
    # Generate for root reports folder
    generate_excel(all_tests, total_tests, passed, failed, pass_rate, f"{OUTPUT_DIR}/reports/E2E_Test_Report.xlsx")
    generate_html(all_tests, total_tests, passed, failed, pass_rate, f"{OUTPUT_DIR}/reports/Test_Summary.html")
    
    # Generate for selenium/reports folder (to match structure exactly)
    generate_excel(all_tests, total_tests, passed, failed, pass_rate, f"{OUTPUT_DIR}/selenium/reports/E2E_Test_Report.xlsx")
    generate_html(all_tests, total_tests, passed, failed, pass_rate, f"{OUTPUT_DIR}/selenium/reports/Test_Summary.html")
    
    # Create zip
    with zipfile.ZipFile(ZIP_NAME, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(OUTPUT_DIR):
            for directory in dirs:
                dir_path = os.path.join(root, directory)
                arcname = os.path.relpath(dir_path, OUTPUT_DIR) + '/'
                zipf.writestr(arcname, '')
            for file in files:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, OUTPUT_DIR)
                zipf.write(file_path, arcname)
                
    print(f"Successfully generated {ZIP_NAME} with identical layout and {total_tests} results.")

if __name__ == "__main__":
    main()
